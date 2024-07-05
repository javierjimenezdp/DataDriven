import openpyxl

class Funexel:
    def __init__(self, driver):
        self.driver = driver

    @staticmethod
    def getRowCount(path, sheetName):
        Workbook = openpyxl.load_workbook(path)
        if sheetName not in Workbook.sheetnames:
            print(f"Hojas disponibles: {Workbook.sheetnames}")
            raise KeyError(f"La hoja '{sheetName}' no existe.")
        sheet = Workbook[sheetName]
        return sheet.max_row

    @staticmethod
    def getColumnCount(file, sheetName):
        Workbook = openpyxl.load_workbook(file)
        if sheetName not in Workbook.sheetnames:
            print(f"Hojas disponibles: {Workbook.sheetnames}")
            raise KeyError(f"La hoja '{sheetName}' no existe.")
        sheet = Workbook[sheetName]
        return sheet.max_column

    @staticmethod
    def readData(path, sheetName, rownum, columno):
        Workbook = openpyxl.load_workbook(path)
        if sheetName not in Workbook.sheetnames:
            print(f"Hojas disponibles: {Workbook.sheetnames}")
            raise KeyError(f"La hoja '{sheetName}' no existe.")
        sheet = Workbook[sheetName]
        return sheet.cell(row=rownum, column=columno).value

    @staticmethod
    def writeData(path, sheetName, rownum, columno, data):
        Workbook = openpyxl.load_workbook(path)
        if sheetName not in Workbook.sheetnames:
            print(f"Hojas disponibles: {Workbook.sheetnames}")
            raise KeyError(f"La hoja '{sheetName}' no existe.")
        sheet = Workbook[sheetName]
        sheet.cell(row=rownum, column=columno).value = data
        Workbook.save(path)
